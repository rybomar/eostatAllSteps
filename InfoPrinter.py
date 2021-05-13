import pathlib
import openpyxl
from openpyxl import Workbook, worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import numpy as np


class InfoPrinter:
    """Class for create xlsx file from numpy array and list of classes."""

    def __init__(self, x_shift=None,
                 y_shift=None,
                 output_file=None,
                 class_names=None,
                 conf_matrix=None
                 ):
        """
        Create object for generate xlsx file from numpy array and list of classes.

        :param x_shift: vertical shift (in x)
        :type  x_shift: int
        :param y_shift: horizontal shift (in y)
        :type y_shift: int
        :param output_file: output file in .xlsx format
        :type output_file: pathlib.Path
        :param class_names: list of class names
        :type class_names: list
        :param conf_matrix: array of classes
        :type conf_matrix: numpy array
        """
        self.x_shift = x_shift
        self.y_shift = y_shift
        self.output_file = output_file
        self.class_names = class_names
        self.conf_matrix = conf_matrix
        # Check the length of class_names and conf_matrix.
        assert len(self.class_names) == self.conf_matrix.shape[0]
        assert len(self.class_names) == self.conf_matrix.shape[1]
        self.array_length = len(self.class_names)
        self.wb = Workbook()

        self.create_xlsx_file_with_default_values()

    # def __del__(self):
    #     self.wb.save(str(self.output_file))

    def __save_output_file(self):
        """
        Save file.

        :return:
        """
        self.wb.save(str(self.output_file))

    @property
    def x_shift(self):
        """
        property x_shift

        :return: x_shift
        :type: int
        """
        return self._x_shift

    @x_shift.setter
    def x_shift(self, value: int = None):
        """
        x_shift setter
        value must be greather then 0

        :param value: x_shift
        :return: none
        :rtype: none
        """
        if value is None:
            value = 0
        assert value >= 0
        self._x_shift = value

    @property
    def y_shift(self):
        """
        property y_shift

        :return: y_shift
        :rtype: int
        """
        return self._y_shift

    @y_shift.setter
    def y_shift(self, value: int = None):
        """
        y_shift setter
        value must be greather then 0

        :param value: y_shift
        :type value: int
        :return: none
        :rtype: none
        """
        if value is None:
            value = 0
        assert value >= 0
        self._y_shift = value

    @property
    def output_file(self):
        """
        property output_file

        :return: output_file
        :rtype: pathlib.Path
        """
        return self._output_file

    @output_file.setter
    def output_file(self, value: pathlib.Path = None):
        """
        output_file setter

        :param value: path to output file
        :type value: pathlib.Path
        :return: none
        :rtype: none
        """
        if value is None:
            value = pathlib.Path("output.xlsx")
        else:
            assert value.suffix == ".xlsx"
        self._output_file = value

    @property
    def class_names(self):
        """
        property class_names

        :return: class_names
        :rtype: list
        """
        return self._class_names

    @class_names.setter
    def class_names(self, value: list = None):
        """
        class_names setter

        :param value: class_names
        :type value: list
        :return: none
        :rtype: none
        """
        if value is None:
            value = ['burak cukrowy', 'gorczyca', 'gryka', 'jęczmień jary', 'jęczmień ozimy', 'kukurydza', 'owies',
                     'plantacje drzew owocowych', 'plantacje krzewów owocowych', 'proso', 'pszenica jara',
                     'pszenica ozima',
                     'pszenżyto jare', 'pszenżyto ozime', 'rzepak jary', 'rzepak ozimy', 'strączkowe', 'truskawka',
                     'TiUZ_MD',
                     'warzywa', 'ziemniaki', 'zioła i przyprawy', 'żyto ozime']
        self._class_names = value

    @property
    def conf_matrix(self):
        """
        property conf_matrix

        :return: conf_matrix
        :rtype: np.ndarray
        """
        return self._conf_matrix

    @conf_matrix.setter
    def conf_matrix(self, value: np.ndarray = None):
        """
        conf_matrix setter
        array must be x dim == y dim

        :param value: conf_matrix
        :type value: np.ndarray
        :return: none
        :rtype: none
        """
        if value is None:
            value = np.array([[207, 0, 0, 0, 1, 1, 0, 0, 0, 3, 0, 0, 0, 0, 1, 0, 0, 0, 0, 2, 2, 0, 0],
                              [0, 61, 14, 4, 1, 5, 12, 6, 15, 10, 4, 0, 5, 0, 0, 1, 9, 10, 9, 0, 1, 5, 2],
                              [0, 1, 88, 0, 0, 8, 3, 2, 6, 6, 0, 0, 3, 1, 1, 0, 8, 8, 6, 3, 6, 2, 7],
                              [0, 1, 2, 153, 0, 1, 14, 0, 0, 1, 9, 1, 4, 0, 0, 0, 1, 0, 0, 0, 0, 2, 0],
                              [0, 2, 0, 3, 154, 0, 0, 0, 2, 0, 1, 4, 1, 3, 0, 1, 2, 0, 1, 0, 0, 0, 0],
                              [1, 0, 2, 1, 2, 183, 0, 0, 2, 0, 0, 2, 2, 0, 1, 0, 2, 0, 1, 3, 1, 1, 2],
                              [0, 5, 0, 5, 3, 2, 140, 0, 1, 1, 11, 0, 9, 1, 0, 0, 2, 0, 2, 1, 0, 1, 3],
                              [0, 0, 0, 0, 0, 0, 0, 145, 29, 1, 0, 0, 1, 0, 0, 0, 1, 6, 2, 1, 0, 0, 0],
                              [0, 4, 3, 0, 0, 2, 1, 12, 94, 6, 0, 0, 1, 2, 0, 0, 0, 15, 15, 0, 1, 0, 0],
                              [0, 3, 13, 1, 1, 9, 2, 2, 3, 134, 0, 1, 0, 1, 3, 0, 0, 1, 0, 2, 0, 0, 0],
                              [0, 2, 0, 9, 0, 1, 20, 0, 1, 2, 91, 22, 12, 5, 1, 0, 10, 2, 0, 1, 2, 2, 1],
                              [0, 0, 0, 1, 3, 0, 1, 0, 0, 0, 5, 190, 2, 12, 0, 1, 0, 0, 0, 1, 0, 0, 2],
                              [0, 4, 0, 5, 1, 2, 21, 0, 1, 0, 18, 0, 50, 46, 0, 0, 10, 0, 2, 0, 1, 0, 11],
                              [0, 0, 0, 1, 5, 0, 0, 0, 1, 1, 0, 3, 10, 171, 0, 0, 0, 0, 1, 0, 0, 0, 12],
                              [0, 8, 0, 0, 0, 3, 0, 1, 1, 0, 1, 1, 0, 2, 71, 60, 2, 1, 1, 1, 0, 0, 2],
                              [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 16, 174, 0, 0, 0, 0, 0, 0, 0],
                              [3, 2, 9, 0, 1, 3, 10, 1, 2, 2, 4, 1, 4, 0, 0, 0, 105, 0, 4, 7, 2, 5, 3],
                              [0, 2, 1, 0, 0, 1, 1, 6, 5, 1, 0, 0, 0, 0, 0, 0, 1, 134, 0, 4, 3, 1, 1],
                              [0, 1, 0, 1, 1, 2, 1, 3, 5, 0, 0, 0, 1, 0, 0, 0, 1, 0, 170, 0, 0, 0, 0],
                              [8, 1, 1, 1, 0, 9, 1, 0, 0, 2, 0, 2, 2, 0, 0, 0, 4, 10, 1, 102, 14, 22, 0],
                              [8, 0, 1, 0, 0, 4, 0, 0, 1, 0, 1, 0, 1, 1, 1, 0, 4, 1, 0, 8, 156, 6, 1],
                              [1, 1, 6, 2, 2, 7, 4, 0, 11, 4, 0, 3, 0, 2, 1, 0, 2, 3, 3, 16, 7, 120, 0],
                              [0, 0, 3, 0, 1, 0, 2, 0, 0, 0, 0, 0, 6, 8, 1, 0, 0, 0, 0, 0, 0, 0, 168]])
        # Check x and y size
        assert value.shape[0] == value.shape[1]
        self._conf_matrix = value

    def __conf_matrix_to_ws(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0, matrix_y_shift: int = 0):
        """
        save conf_matrix to ws

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        for column_number in range(self.array_length):
            for row_number in range(self.array_length):
                ws[get_column_letter(column_number + 1 + self.x_shift + matrix_x_shift) + str(
                    row_number + 1 + self.y_shift + matrix_y_shift)].value = \
                    int(self.conf_matrix[column_number][row_number])

    def __vertical_legend_class_names_to_ws(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                                            matrix_y_shift: int = 0):
        """
        save vertical legend class names to ws

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        for y in range(self.array_length):
            ws[get_column_letter(self.x_shift + matrix_x_shift) + str(
                y + 1 + self.y_shift + matrix_y_shift)].value = str(
                self.class_names[y])

    def __vertical_numbers_class_names_to_ws(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                                             matrix_y_shift: int = 0):
        """
        calculate vertical number class names to ws

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        for y in range(self.array_length):
            ws[get_column_letter(self.x_shift - 1 + matrix_x_shift) +
               str(y + 1 + self.y_shift + matrix_y_shift)].value = str(y + 1)

    def __horizontal_legend_class_names_to_ws(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                                              matrix_y_shift: int = 0):
        """
        calculate horizontal legend class names to ws

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift must be greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift must be greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        for x in range(self.array_length):
            ws[get_column_letter(x + 1 + self.x_shift + matrix_x_shift) + str(
                -1 + 1 + self.y_shift + matrix_y_shift)].value = str(
                self.class_names[x])

    def __horizontal_numbers_class_names_to_ws(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                                               matrix_y_shift: int = 0):
        """
        calculate horizontal numbers class names to ws

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift must be greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift must be greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        for x in range(self.array_length):
            ws[get_column_letter(x + 1 + self.x_shift + matrix_x_shift) +
               str(-2 + 1 + self.y_shift + matrix_y_shift)].value = str(x + 1)

    def __horizontal_sum_of_all(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0, matrix_y_shift: int = 0):
        """
        calculate horizontal sum of all

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        for column_number in range(self.array_length):
            formula = "=sum(" + get_column_letter(column_number + 1 + self.x_shift + matrix_x_shift) + str(
                1 + self.y_shift + matrix_y_shift) + ":" + get_column_letter(
                column_number + 1 + self.x_shift + matrix_x_shift) + str(
                self.y_shift + matrix_y_shift + self.array_length) + str(")")
            ws[get_column_letter(column_number + 1 + self.x_shift + matrix_x_shift) + str(
                self.array_length + 1 + self.y_shift + matrix_y_shift)].value = formula

    def __horizontal_percent_of_all(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0, matrix_y_shift: int = 0,
                                    number_format: str = '0.00%'):
        """
        calculate horizontal percent of all

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :param number_format: number format default 0.00%
        :type number_format: str
        :return: none
        :rtype: none
        """
        row_number: int = 0
        for column_number in range(self.array_length):
            formula = "=" + get_column_letter(column_number + 1 + self.x_shift + matrix_x_shift) + str(
                1 + row_number + self.y_shift + matrix_y_shift) + "/" + get_column_letter(
                column_number + 1 + self.x_shift + matrix_x_shift) \
                      + str(1 + self.y_shift + matrix_y_shift + self.array_length)
            ws[get_column_letter(column_number + 1 + self.x_shift + matrix_x_shift) + str(
                self.array_length + 2 + self.y_shift + matrix_y_shift)].value = formula
            ws[get_column_letter(column_number + 1 + self.x_shift + matrix_x_shift) + str(
                self.array_length + 2 + self.y_shift + matrix_y_shift)].number_format = number_format
            row_number += 1

    def __vertical_sum_of_all(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0, matrix_y_shift: int = 0):
        """
        calculate vertical sum of all

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        for row_number in range(self.array_length):
            formula = "=sum(" + get_column_letter(self.x_shift + 1 + matrix_x_shift) + str(
                1 + row_number + self.y_shift + matrix_y_shift) + ":" + \
                      get_column_letter(self.array_length + self.x_shift + matrix_x_shift) + \
                      str(row_number + 1 + self.y_shift + matrix_y_shift) + str(")")
            ws[get_column_letter(self.x_shift + 1 + matrix_x_shift + self.array_length) +
               str(row_number + 1 + self.y_shift + matrix_y_shift)].value = formula

    def __vertical_percent_of_all(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0, matrix_y_shift: int = 0,
                                  number_format: str = '0.00%'):
        """
        calculate vertical percent of all

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :param number_format: number format default 0.00%
        :type number_format: str
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        column_number: int = 0
        for row_number in range(self.array_length):
            formula = "=" + get_column_letter(1 + row_number + self.x_shift + matrix_x_shift) + str(
                1 + column_number + self.y_shift + matrix_y_shift) + "/" + \
                      get_column_letter(1 + self.array_length + self.x_shift + matrix_x_shift) \
                      + str(1 + row_number + self.y_shift + matrix_y_shift)
            ws[get_column_letter(2 + self.array_length + self.x_shift + matrix_x_shift) +
               str(row_number + 1 + self.y_shift + matrix_y_shift)].value = formula
            ws[get_column_letter(2 + self.array_length + self.x_shift + matrix_x_shift) +
               str(row_number + 1 + self.y_shift + matrix_y_shift)].number_format = number_format
            column_number += 1

    def __set_title(self, ws: openpyxl.worksheet, sheet_title: str, matrix_x_shift: int = 0, matrix_y_shift: int = 0):
        """
        set title to spreadsheet

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param sheet_title: spreadsheet title
        :type sheet_title str
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        if matrix_x_shift > 0 or matrix_y_shift > 0:
            ws.merge_cells(get_column_letter(1 + self.x_shift) + str(1 + self.y_shift) + ':' + get_column_letter(
                matrix_x_shift + self.x_shift) + str(matrix_y_shift + self.y_shift))
        ws[get_column_letter(1 + self.x_shift) + str(1 + self.y_shift)].value = sheet_title
        ws[get_column_letter(1 + self.x_shift) + str(1 + self.y_shift)].alignment = Alignment(vertical='center',
                                                                                              horizontal='center',
                                                                                              text_rotation=0,
                                                                                              wrap_text=True)

    def __calculate_over_all_accuracy(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                                      matrix_y_shift: int = 0, number_format: str = '0.00%'):
        """
        calculate over all accuracy

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :param number_format: number format default 0.00
        :type number_format: str
        :return: none
        :rtype: none
        """

        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        x: int = 1 + self.x_shift + self.array_length + matrix_x_shift
        y: int = 2 + self.y_shift + self.array_length + matrix_y_shift
        formula: str = '=('
        for matrix_diagonal in range(self.array_length):
            formula += get_column_letter(matrix_diagonal + 1 + self.x_shift + matrix_x_shift) + str(
                matrix_diagonal + 1 + self.y_shift + matrix_y_shift)
            if matrix_diagonal != self.array_length - 1:
                formula += '+'
        formula += ')/' + get_column_letter(1 + self.x_shift + self.array_length + matrix_x_shift) + str(
            1 + self.y_shift + self.array_length + matrix_y_shift)
        ws[get_column_letter(x) + str(y)].value = 'OA'
        ws[get_column_letter(x) + str(y)].alignment = Alignment(vertical='center', horizontal='center')
        ws[get_column_letter(x + 1) + str(y)].value = formula
        ws[get_column_letter(x + 1) + str(y)].number_format = number_format

    def __calculate_sum_diag(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                             matrix_y_shift: int = 0):
        """
        calculate sum diagonal

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        x: int = 1 + self.x_shift + self.array_length + matrix_x_shift
        y: int = 3 + self.y_shift + self.array_length + matrix_y_shift
        formula: str = '='
        for matrix_diagonal in range(self.array_length):
            formula += get_column_letter(matrix_diagonal + 1 + self.x_shift + matrix_x_shift) + str(
                matrix_diagonal + 1 + self.y_shift + matrix_y_shift)
            if matrix_diagonal != self.array_length - 1:
                formula += '+'
        ws[get_column_letter(x) + str(y)].value = u'\u03A3' + " diag"
        ws[get_column_letter(x) + str(y)].alignment = Alignment(vertical='center', horizontal='left')
        ws[get_column_letter(x + 1) + str(y)].value = formula

    def __calculate_sum_x_ij_multiplication_y_ij(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                                                 matrix_y_shift: int = 0):
        """
        calculate sun x ij * y ij

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        formula = '='
        for matrix_diagonal in range(self.array_length):
            formula += get_column_letter(matrix_diagonal + 1 + self.x_shift + matrix_x_shift) + str(
                self.array_length + 1 + self.y_shift + matrix_y_shift) + '*' + \
                       get_column_letter(self.x_shift + 1 + matrix_x_shift + self.array_length) + \
                       str(matrix_diagonal + 1 + self.y_shift + matrix_y_shift)
            if matrix_diagonal != self.array_length - 1:
                formula += '+'
        x: int = 1 + self.x_shift + self.array_length + matrix_x_shift
        y: int = 4 + self.y_shift + self.array_length + matrix_y_shift
        ws[get_column_letter(x) + str(y)].value = u'\u03A3' + "Xij*Yij"
        ws[get_column_letter(x) + str(y)].alignment = Alignment(vertical='center', horizontal='left')
        ws[get_column_letter(x + 1) + str(y)].value = formula

    def __calculate_kappa(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                          matrix_y_shift: int = 0, number_format: str = '0.00'):
        """
        calculate kappa

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :param number_format: number format default 0.00
        :type number_format: str
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        x: int = 1 + self.x_shift + self.array_length + matrix_x_shift
        y: int = 5 + self.y_shift + self.array_length + matrix_y_shift
        vertical_sum = get_column_letter(x) + str(y - 4)
        sum_diag = get_column_letter(x + 1) + str(y - 2)
        sum_x_ij_y_ij = get_column_letter(x + 1) + str(y - 1)
        formula: str = '=(' + vertical_sum + '*' + sum_diag + '-' + sum_x_ij_y_ij + ')' + '/' + '(' + vertical_sum + \
                       '^2' + '-' + sum_x_ij_y_ij + ')'
        ws[get_column_letter(x) + str(y)].value = "Kappa"
        ws[get_column_letter(x) + str(y)].alignment = Alignment(vertical='center', horizontal='left')
        ws[get_column_letter(x + 1) + str(y)].value = formula
        ws[get_column_letter(x + 1) + str(y)].number_format = number_format

    def __calculate_vertical_sum(self, ws: openpyxl.worksheet, matrix_x_shift: int = 0,
                                 matrix_y_shift: int = 0):
        """
        calculate vertical sum

        :param ws: worksheet
        :type ws: openpyxl.worksheet
        :param matrix_x_shift: matrix x shift greater then 0
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 0
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 0
        assert matrix_y_shift >= 0
        formula = "=sum(" + get_column_letter(self.x_shift + matrix_x_shift + self.array_length + 1) + \
                  str(self.y_shift + matrix_y_shift) + ':' + \
                  get_column_letter(self.x_shift + matrix_x_shift + self.array_length + 1) + \
                  str(self.y_shift + matrix_y_shift + self.array_length) + ')'
        ws[get_column_letter(1 + self.x_shift + self.array_length + matrix_x_shift) +
           str(1 + self.y_shift + self.array_length + matrix_y_shift)].value = formula

    def create_xlsx_file_with_default_values(self, sheet_name: str = "Classify",
                                             sheet_title: str = "Classification table",
                                             matrix_x_shift: int = 2, matrix_y_shift: int = 2):
        """
        create xlsx file with default values

        :param sheet_name: name of spreadsheet
        :type sheet_name: str
        :param sheet_title: spreadsheet title
        :type sheet_title: str
        :param matrix_x_shift: matrix x shift greater then 2
        :type matrix_x_shift: int
        :param matrix_y_shift: matrix y shift greater then 2
        :type matrix_y_shift: int
        :return: none
        :rtype: none
        """
        assert matrix_x_shift >= 2
        assert matrix_y_shift >= 2
        self.conf_matrix = self.conf_matrix.transpose()
        ws = self.wb.active
        ws.title = sheet_name
        self.__conf_matrix_to_ws(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__vertical_legend_class_names_to_ws(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__vertical_numbers_class_names_to_ws(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__horizontal_legend_class_names_to_ws(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__horizontal_numbers_class_names_to_ws(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__horizontal_sum_of_all(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__horizontal_percent_of_all(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__vertical_sum_of_all(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__vertical_percent_of_all(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__set_title(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws, sheet_title=sheet_title)
        self.__calculate_over_all_accuracy(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__calculate_sum_diag(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__calculate_sum_x_ij_multiplication_y_ij(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift,
                                                      ws=ws)
        self.__calculate_kappa(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__calculate_vertical_sum(matrix_x_shift=matrix_x_shift, matrix_y_shift=matrix_y_shift, ws=ws)
        self.__save_output_file()


# output = InfoPrinter()
# output.create_xlsx_file_with_default_values()
